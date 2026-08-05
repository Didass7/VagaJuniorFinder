import os
import datetime
from typing import List, Optional, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import CandidateProfile
from matcher import ScoredJob
from scraper import Job

class ExcelReportBuilder:
    """Builds beautiful, interactive Excel spreadsheets (.xlsx) for daily job reports and master database tracking."""
    def __init__(self, profile: CandidateProfile):
        self.profile = profile

    def _create_styled_worksheet(self, wb: openpyxl.Workbook, sheet_title: str, scored_jobs: List[ScoredJob], is_master: bool = False):
        ws = wb.create_sheet(title=sheet_title)
        
        # Headers definition
        headers = [
            "Match Score (%)",
            "Título da Vaga",
            "Empresa",
            "Localização",
            "Modo de Trabalho",
            "Estágio IEFP?",
            "Tecnologias Encontradas",
            "Nível Senioridade",
            "Data Publicação",
            "Fonte",
            "Link de Candidatura",
            "Estado Candidatura",
            "Notas Pessoais"
        ]
        
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sort jobs by match score descending
        sorted_jobs = sorted(scored_jobs, key=lambda j: j.score, reverse=True)

        # Append data rows
        for row_idx, sj in enumerate(sorted_jobs, start=2):
            job = sj.job
            skills_str = ", ".join(sj.matched_skills) if sj.matched_skills else "N/A"
            iefp_str = "Sim (IEFP / ATIVAR)" if job.iefp_mentioned else "Não especificado"
            
            ws.cell(row=row_idx, column=1, value=round(sj.score, 1))
            ws.cell(row=row_idx, column=2, value=job.title)
            ws.cell(row=row_idx, column=3, value=job.company)
            ws.cell(row=row_idx, column=4, value=job.location)
            ws.cell(row=row_idx, column=5, value=job.work_mode)
            ws.cell(row=row_idx, column=6, value=iefp_str)
            ws.cell(row=row_idx, column=7, value=skills_str)
            ws.cell(row=row_idx, column=8, value=sj.seniority_status)
            ws.cell(row=row_idx, column=9, value=job.pub_date)
            ws.cell(row=row_idx, column=10, value=job.source)
            
            # Clickable hyperlink in Excel
            link_cell = ws.cell(row=row_idx, column=11, value="Abrir Vaga 🔗")
            link_cell.hyperlink = job.link
            link_cell.font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
            
            # Interactive columns for tracking candidate progress
            ws.cell(row=row_idx, column=12, value="Por Candidatar")
            ws.cell(row=row_idx, column=13, value="")

            # Styling data cells
            for col_num in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                cell.font = Font(name="Segoe UI", size=10)
                
                # Center-align specific columns
                if col_num in [1, 5, 6, 8, 9, 10, 11, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Score color highlights
                if col_num == 1:
                    score_val = sj.score
                    if score_val >= 75.0:
                        cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")  # Soft green
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="375623")
                    elif score_val >= 55.0:
                        cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")  # Soft yellow
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="7F6000")

        # Add Data Validation dropdown for "Estado Candidatura" (Column L / 12)
        dv = DataValidation(
            type="list",
            formula1='"Por Candidatar,Candidatado,Entrevista,Proposta,Rejeitado,Sem Resposta"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        max_row = max(len(sorted_jobs) + 1, 2)
        dv.add(f"L2:L{max_row}")

        # Set row height
        ws.row_dimensions[1].height = 28
        for r in range(2, max_row + 1):
            ws.row_dimensions[r].height = 22

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        # Freeze header row
        ws.freeze_panes = "A2"

    def build_daily_report(self, scored_jobs: List[ScoredJob], output_dir: str = "reports") -> str:
        """Generates the daily Excel report with tabs for Top Matches, Promising Matches, and All Ingested Jobs."""
        os.makedirs(output_dir, exist_ok=True)
        date_str = datetime.date.today().strftime("%Y-%m-%d")
        filepath = os.path.join(output_dir, f"jobs_report_{date_str}.xlsx")

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        top_matches = [j for j in scored_jobs if j.score >= 75.0]
        promising_matches = [j for j in scored_jobs if 55.0 <= j.score < 75.0]

        # Tab 1: Top Matches
        self._create_styled_worksheet(wb, "🎯 Top Vagas (≥75%)", top_matches)
        
        # Tab 2: Promising Matches
        self._create_styled_worksheet(wb, "💡 Vagas Promissoras (55-74%)", promising_matches)
        
        # Tab 3: All Scraped & Evaluated Jobs
        self._create_styled_worksheet(wb, "🗂️ Todas as Vagas Evaluated", scored_jobs)

        try:
            wb.save(filepath)
        except PermissionError:
            alt_filepath = filepath.replace(".xlsx", "_latest.xlsx")
            wb.save(alt_filepath)
            return alt_filepath
        return filepath

    def update_master_database(self, scored_jobs: List[ScoredJob], master_filepath: str = os.path.join("data", "jobs_database.xlsx")) -> str:
        """Maintains a cumulative master Excel database accumulating all jobs ever found and preserving candidate notes."""
        master_jobs_dict: Dict[str, ScoredJob] = {}
        user_notes_dict: Dict[str, Dict[str, str]] = {}

        # Fallback to root jobs_database.xlsx if master_filepath doesn't exist yet
        if not os.path.exists(master_filepath) and os.path.exists("jobs_database.xlsx"):
            master_filepath_to_read = "jobs_database.xlsx"
        else:
            master_filepath_to_read = master_filepath

        # 1. Read existing rows from master_filepath if it exists to preserve cumulative history
        if os.path.exists(master_filepath_to_read):
            try:
                from matcher import JobMatcher
                matcher = JobMatcher(profile=self.profile)

                old_wb = openpyxl.load_workbook(master_filepath_to_read)
                if "🗃️ Base de Dados Master" in old_wb.sheetnames:
                    old_ws = old_wb["🗃️ Base de Dados Master"]
                    for row in old_ws.iter_rows(min_row=2, values_only=False):
                        if row[10] and (row[10].hyperlink or row[10].value):
                            url = row[10].hyperlink.target if row[10].hyperlink else str(row[10].value or "")
                            if not url:
                                continue
                            
                            try:
                                score = float(row[0].value or 0.0)
                            except ValueError:
                                score = 0.0

                            title = str(row[1].value or "")
                            company = str(row[2].value or "")
                            location = str(row[3].value or "")
                            work_mode = str(row[4].value or "")
                            iefp_str = str(row[5].value or "")
                            skills_str = str(row[6].value or "")
                            seniority = str(row[7].value or "")
                            pub_date = str(row[8].value or "")
                            source = str(row[9].value or "")
                            status = str(row[11].value or "Por Candidatar")
                            notes = str(row[12].value or "")
                            
                            user_notes_dict[url] = {"status": status, "notes": notes}
                            
                            reconstructed_job = Job(
                                title=title, company=company, location=location,
                                work_mode=work_mode, link=url,
                                description=f"{title} na empresa {company} em {location}.",
                                source=source, pub_date=pub_date
                            )
                            if "Sim" in iefp_str:
                                reconstructed_job.iefp_mentioned = True

                            # Check if title or seniority indicates senior/leadership/teaching disqualification
                            evaluated = matcher.evaluate_job(reconstructed_job)
                            if evaluated.score == 0.0 or "Sénior" in evaluated.seniority_status or "Liderança" in evaluated.seniority_status or "Irrelevante" in evaluated.seniority_status:
                                # Purge disqualified job from master database
                                continue

                            skills_list = [s.strip() for s in skills_str.split(",") if s.strip() and s.strip() != "N/A"]
                            master_jobs_dict[url] = ScoredJob(
                                job=reconstructed_job, score=score, matched_skills=skills_list,
                                missing_skills=[], seniority_status=seniority
                            )
            except Exception:
                pass

        # 2. Merge new scored_jobs into master_jobs_dict
        for sj in scored_jobs:
            url = sj.job.link
            # If job is disqualified (score == 0.0 or seniority disqualification), remove it from master database
            if sj.score == 0.0 or "Sénior" in sj.seniority_status or "Experiência >" in sj.seniority_status or "Irrelevante" in sj.seniority_status or "Incompleta" in sj.seniority_status:
                if url in master_jobs_dict:
                    del master_jobs_dict[url]
            else:
                master_jobs_dict[url] = sj

        combined_jobs = list(master_jobs_dict.values())

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Build master sheet with all cumulative jobs
        self._create_styled_worksheet(wb, "🗃️ Base de Dados Master", combined_jobs, is_master=True)
        ws = wb["🗃️ Base de Dados Master"]

        # Restore candidate tracking statuses and notes
        for row_idx in range(2, ws.max_row + 1):
            cell_link = ws.cell(row=row_idx, column=11)
            if cell_link and cell_link.hyperlink:
                url = cell_link.hyperlink.target
                if url in user_notes_dict:
                    ws.cell(row=row_idx, column=12, value=user_notes_dict[url]["status"])
                    ws.cell(row=row_idx, column=13, value=user_notes_dict[url]["notes"])

        dir_name = os.path.dirname(master_filepath)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        try:
            wb.save(master_filepath)
        except PermissionError:
            alt_filepath = "jobs_database_latest.xlsx"
            wb.save(alt_filepath)
            return alt_filepath

        return master_filepath
